import sys
import openpyxl
from PySide2.QtWidgets import *
from PySide2.QtCore import QStringListModel
from UI import Ui_MainWindow
import sqlite3

workbook = openpyxl.Workbook()
# 制表AirportCommunication 0
sheetAirportCommunication = workbook.create_sheet("AirportCommunication",0)
sheetAirportCommunication.cell(1,1).value='area_code'
sheetAirportCommunication.cell(1,2).value='icao_code'
sheetAirportCommunication.cell(1,3).value='airport_identifier'
sheetAirportCommunication.cell(1,4).value='communication_type'
sheetAirportCommunication.cell(1,5).value='communication_frequency'
sheetAirportCommunication.cell(1,6).value='frequency_units'
sheetAirportCommunication.cell(1,7).value='service_indicator'
sheetAirportCommunication.cell(1,8).value='callsign'
sheetAirportCommunication.cell(1,9).value='latitude'
sheetAirportCommunication.cell(1,10).value='longtitude'
# 制表AirportCommunication 1
sheetAirportLookup = workbook.create_sheet("AirportLookup",1)
sheetAirportLookup.cell(1, 1).value='extID'
sheetAirportLookup.cell(1, 2).value='ID'
# 制表Airports 2
sheetAirports = workbook.create_sheet("Airports",2)
sheetAirports.cell(1, 1).value='ID'
sheetAirports.cell(1, 2).value='Name'
sheetAirports.cell(1, 3).value='ICAO'
sheetAirports.cell(1, 4).value='PrimaryID'
sheetAirports.cell(1,5).value='Latitude'
sheetAirports.cell(1, 6).value='Longtitude'
sheetAirports.cell(1, 7).value='Elevation'
sheetAirports.cell(1, 8).value='TransitionAltitude'
sheetAirports.cell(1, 9).value='TransitionLevel'
sheetAirports.cell(1, 10).value='SpeedLimit'
sheetAirports.cell(1, 11).value= 'SpeedLimitAltitude'
# 制表Airways 3
sheetAirways = workbook.create_sheet("Airways",3)
# 制表config 4
sheetconfig = workbook.create_sheet("config",4)
# 制表Gls 5
sheetGls = workbook.create_sheet("Gls",5)
# 制表GridMora 6
sheetGridMora = workbook.create_sheet("GridMora",6)
# 制表Holdings 7
sheetHoldings = workbook.create_sheet("Holdings",7)
# 制表ILSes 8
sheetILSes = workbook.create_sheet("ILSes",8)
sheetILSes.cell(1, 1).value='ID'
sheetILSes.cell(1, 2).value='RunwayID'
sheetILSes.cell(1, 3).value='Freq'
sheetILSes.cell(1, 4).value='GsAngle'
sheetILSes.cell(1, 5).value='Latitude'
sheetILSes.cell(1, 6).value='Longtitude'
sheetILSes.cell(1, 7).value='Category'
sheetILSes.cell(1, 8).value='Ident'
sheetILSes.cell(1, 9).value='LocCourse'
sheetILSes.cell(1, 10).value='CrossingHeight'
sheetILSes.cell(1, 11).value='HasDme'
sheetILSes.cell(1, 12).value='Elevation'
# 制表Markers 9
sheetMarkers = workbook.create_sheet("Markers",9)
# 制表MarkerTypes 10
sheetMarkerTypes = workbook.create_sheet("MarkerTypes",10)
# 制表NavaidLookup 11
sheetNavaidLookup = workbook.create_sheet("NavaidLookup",11)
# 制表Navaids 12
sheetNavaids = workbook.create_sheet("Navaids",12)
# 制表NavaidTypes 13
sheetNavaidTypes = workbook.create_sheet("NavaidTypes",13)
# 制表Runways 14
sheetRunways = workbook.create_sheet("Runways",14)
sheetRunways.cell(1, 1).value='ID'
sheetRunways.cell(1, 2).value='AirportID'
sheetRunways.cell(1, 3).value='Ident'
sheetRunways.cell(1, 4).value='TrueHeading'
sheetRunways.cell(1, 5).value='Length'
sheetRunways.cell(1, 6).value='Width'
sheetRunways.cell(1, 7).value='Surface'
sheetRunways.cell(1, 8).value='Latitude'
sheetRunways.cell(1, 9).value='Longtitude'
sheetRunways.cell(1, 10).value='Elevation'
# 制表SurfaceTypes 15
sheetSurfaceTypes = workbook.create_sheet("SurfaceTypes",15)
# 制表TerminalLegs 16
sheetTerminalLegs = workbook.create_sheet("TerminalLegs",16)
# 制表TerminalLegsEx 17
sheetTerminalLegsEx = workbook.create_sheet("TerminalLegsEx",17)
# 制表Terminals 18
sheetTerminals = workbook.create_sheet("Terminals",18)
# 制表TrmLegTypes 19
sheetTrmLegTypes = workbook.create_sheet("TrmLegTypes",19)
# 制表WaypointLookup 20
sheetWaypointLookup = workbook.create_sheet("WaypointLookup",20)
sheetWaypointLookup.cell(1,1).value = 'Ident'
sheetWaypointLookup.cell(1,2).value = 'Country'
sheetWaypointLookup.cell(1,3).value = 'ID'
# 制表Waypoints 21
sheetWaypoints = workbook.create_sheet("Waypoints",21)
sheetWaypoints.cell(1,1).value = 'ID'
sheetWaypoints.cell(1,2).value = 'Ident'
sheetWaypoints.cell(1,3).value = 'Collocated'
sheetWaypoints.cell(1,4).value = 'Name'
sheetWaypoints.cell(1,5).value = 'Latitude'
sheetWaypoints.cell(1,6).value = 'Longtitude'
sheetWaypoints.cell(1,7).value = 'NavaidID'
# Airways 22
sheetAirwayLegs = workbook.create_sheet("AirwayLegs",22)
curRow = [2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2]


class MyMainForm(QMainWindow, Ui_MainWindow):

    def cor2lat(self,lat):

        NE = False
        if float(lat)<0:
            lat = (-float(lat)).__str__()
            NE =True
        list = lat.__str__().split(".")
        lengg = len(list[0])
        for i in range(2 - lengg):
            list[0] = '0' + list[0]
        temp1 = float("0."+list[1]) * 60
        list2 = temp1.__str__().split(".")
        temp2 = float("0."+list2[1])*60
        if len(list2[0])==0:
            list2[0] = '00'+list2[0]
        else :
            if len(list2[0])==1:
                list2[0] = '0'+list2[0]
        list3 = temp2.__str__().split(".")
        if len(list3[0])==0:
            list2[0] = '0'+list3[0]
        if len(list3[0])==1:
            list3[0] = '0'+list3[0]
        if NE :
            list[0] = 'S'+list[0]
        else:
            list[0] = 'N' + list[0]
        if list3[1]=="0":
            trueLat = list[0] + list2[0][0:2]+list3[0]
        else:
            if len(list3[1]) > 1:
                if int(int(list3[1][0:1]) + int(list3[1][1:2]) / 5) == 0:
                    trueLat = list[0] + list2[0][0:2] + list3[0]
                else:
                    if int(int(list3[1][0:1]) + int(list3[1][1:2]) / 5) == 10:
                        trueLat = list[0] + list2[0][0:2] + (int(list3[0]) + 1).__str__()
                    else:
                        trueLat = list[0] + list2[0][0:2] + list3[0] + "." + int(int(list3[1][0:1]) + int(list3[1][1:2]) / 5).__str__()
            else:
                print(list3[1])
                if len(list3[1]) == 1 and list3[1] != "0":
                    trueLat = list[0] + list2[0][0:2] + list3[0] + "." + int(list3[1][0:1]).__str__()
                else:

                    if list3[1] == "0":
                        trueLat = list[0] + list2[0][0:2] + list3[0]
                    else:
                        trueLat = list[0] + list2[0][0:2] + (int(list3[0]) + 1).__str__()

        return trueLat

    def cor2longt(self,longt):
        NE = False
        if float(longt) < 0:
            longt = (-float(longt)).__str__()
            NE = True
        list = longt.__str__().split(".")
        lengg = len(list[0])
        for i in range(2 - lengg):
            list[0] = '0' + list[0]
        temp1 = float("0." + list[1]) * 60
        list2 = temp1.__str__().split(".")
        temp2 = float("0." + list2[1]) * 60
        if len(list2[0]) == 0:
            list2[0] = '00' + list2[0]
        else:
            if len(list2[0]) == 1:
                list2[0] = '0' + list2[0]
        list3 = temp2.__str__().split(".")

        if len(list3[0]) == 0:
            list2[0] = '0' + list3[0]
        if len(list3[0]) == 1:
            list3[0] = '0' + list3[0]
        if NE :
            list[0] = 'W'+list[0]
        else:
            list[0] = 'E' + list[0]

        if list3[1]=="0":
            trueLong = list[0] + list2[0][0:2]+list3[0]
        else:
            if len(list3[1]) > 1:
                if int(int(list3[1][0:1]) + int(list3[1][1:2]) / 5) == 0:
                    trueLong = list[0] + list2[0][0:2] + list3[0]
                else:
                    if int(int(list3[1][0:1]) + int(list3[1][1:2]) / 5) == 10:
                        trueLong = list[0] + list2[0][0:2] + (int(list3[0]) + 1).__str__()
                    else:
                        trueLong = list[0] + list2[0][0:2] + list3[0] + "." + int(
                            int(list3[1][0:1]) + int(list3[1][1:2]) / 5).__str__()
            else:
                print(list3[1])
                if len(list3[1]) == 1 and list3[1] != "0":
                    trueLong = list[0] + list2[0][0:2] + list3[0] + "." + int(list3[1][0:1]).__str__()
                else:

                    if list3[1] == "0":
                        trueLong = list[0] + list2[0][0:2] + list3[0]
                    else:
                        trueLong = list[0] + list2[0][0:2] + (int(list3[0]) + 1).__str__()
        return trueLong

    def longt2cor(self,longt):
        trueLongt = (int(longt[1:4]) + float(longt[4:6]) / 60.0 + float(longt[6:]) / 3600.0).__str__()
        if longt[0] =="W":
            trueLongt = "-"+trueLongt
        return trueLongt

    def lat2cor(self, lat):
        trueLat = (int(lat[1:3]) + float(lat[3:5]) / 60.0 + float(lat[5:]) / 3600.0).__str__()
        if lat[0]=="S":
            trueLat = "-"+trueLat
        return trueLat

    def __init__(self, parent=None):
        super(MyMainForm, self).__init__(parent)
        self.setupUi(self)
        self.OpenDB.triggered.connect(self.connectdb)
        self.SaveRunway.clicked.connect(self.saverunway)
        self.SaveILS.clicked.connect(self.saveils)

    def connectdb(self):

        openfile_name = QFileDialog.getOpenFileNames(self, '选择文件', filter="Fenix导航数据文件 (nd.db3)",
                                                     dir='C:\\ProgramData\\Fenix\\Navdata')
        if openfile_name[0] != None:
            self.con = sqlite3.connect(openfile_name[0][0])
            self.cur = self.con.cursor()
            self.cur.execute("SELECT ID,ICAO FROM Airports")
            airport = self.cur.fetchall()
            # 实例化列表模型，添加数据
            slm = QStringListModel()
            self.AirportList = []
            for per in airport:
                self.AirportList.append(per[1])

            # 设置模型列表视图，加载数据列表
            slm.setStringList(self.AirportList)

            # 设置列表视图的模型
            self.AirportSelect.setModel(slm)
            # 单击触发自定义的槽函数
            self.AirportSelect.clicked.connect(self.selectairport)
            self.SaveAirport.clicked.connect(self.saveairport)
            self.SaveExcel.triggered.connect(self.saveexcel)
            self.AirportSearchButton.clicked.connect(self.searchAirport)
            self.WPSearchButton.clicked.connect(self.searchwp)
            self.WPSave.clicked.connect(self.savewp)
            self.TMAAPSearchButton.clicked.connect(self.searchtma)
            self.cur.execute("SELECT SurfaceType FROM SurfaceTypes")
            data = self.cur.fetchall()
            self.RSurface.clear()
            for per in data:
                self.RSurface.addItem(per[0].__str__())
            self.ILSCategory.addItem('I')
            self.ILSCategory.addItem('II')
            self.ILSCategory.addItem('III')

    def searchtma(self):
        self.cur.execute("SELECT * FROM Terminals WHERE ICAO = \'"+self.TMAAPSearch.text()+"\'")
        data = self.cur.fetchall()

        self.tmaListView = QStringListModel()
        self.tmaList = []
        for per in data:
            self.tmaList.append(per[4])
        self.label_25.setText("当前机场："+self.TMAAPSearch.text())
        # 设置模型列表视图，加载数据列表
        self.tmaListView.setStringList(self.tmaList)

        # 设置列表视图的模型
        self.TMAList.setModel(self.tmaListView)

        self.TMAList.clicked.connect(self.findline)

    def findline(self):
        self.cur.execute("SELECT ID FROM Terminals WHERE FullName= \'" + self.TMAList.selectedIndexes()[0].data().__str__() + "\'")
        data = self.cur.fetchall()
        print(data[0][0])
        TMAID = data[0][0]
        self.cur.execute("SELECT * FROM TerminalLegs WHERE TerminalID= \'" + TMAID.__str__() + "\'")
        data = self.cur.fetchall()
        strA = ''
        for per in data:
            print(per[5].__str__())

        self.linesView = QStringListModel()
        self.lines = []
        for per in data:
            self.lines.append(per[0].__str__())

        # 设置模型列表视图，加载数据列表
        self.linesView.setStringList(self.lines)

        # 设置列表视图的模型
        self.LineList.setModel(self.linesView)
        self.LineList.clicked.connect(self.selectline)

    def selectline(self):
        #print(self.LineList.selectedIndexes()[0].data())
        self.cur.execute( "SELECT * FROM TerminalLegs WHERE ID = \'" + self.LineList.selectedIndexes()[0].data() + "\'")
        data = self.cur.fetchall()
        print(data)

    def savewp(self):

        global sheetWaypoints
        global sheetWaypointLookup
        maxr = sheetWaypointLookup.max_row+1
        for i in range(2,maxr+1):
            if sheetWaypointLookup.cell(i,3).value.__str__() == self.WPID.text().__str__():
                print('发现重合,在第'+i.__str__()+'行')
                sheetWaypointLookup.cell(i, 1).value = self.WPIdent.text()
                sheetWaypointLookup.cell(i, 2).value = self.WPCountry.text()
                sheetWaypointLookup.cell(i, 3).value = self.WPID.text()
                break
            if i == maxr:
                print('没有重合,在第'+i.__str__()+'行')
                sheetWaypointLookup.cell(i, 1).value = self.WPIdent.text()
                sheetWaypointLookup.cell(i, 2).value = self.WPCountry.text()
                sheetWaypointLookup.cell(i, 3).value = self.WPID.text()
        maxr = sheetWaypoints.max_row+1
        for i in range(2,maxr+1):
            if sheetWaypoints.cell(i,1).value.__str__() == self.WPID.text().__str__():
                print('发现重合,在第'+i.__str__()+'行')
                sheetWaypoints.cell(i, 1).value = self.WPID.text()
                sheetWaypoints.cell(i, 2).value = self.WPIdent.text()
                sheetWaypoints.cell(i, 3).value = self.WPCollcated.text()
                sheetWaypoints.cell(i, 4).value = self.WPName.text()
                sheetWaypoints.cell(i, 5).value = self.lat2cor(self.WPLat.text())
                sheetWaypoints.cell(i, 5).number_format = '-#.######'
                sheetWaypoints.cell(i, 6).value = self.longt2cor(self.WPLongt.text())
                sheetWaypoints.cell(i, 6).number_format = '-#.######'
                sheetWaypoints.cell(i, 7).value = self.WPNav.text()
                break
            if i == maxr:
                print('没有重合,在第' + i.__str__() + '行')
                sheetWaypoints.cell(i, 1).value = self.WPID.text()
                sheetWaypoints.cell(i, 2).value = self.WPIdent.text()
                sheetWaypoints.cell(i, 3).value = self.WPCollcated.text()
                sheetWaypoints.cell(i, 4).value = self.WPName.text()
                sheetWaypoints.cell(i, 5).value = self.lat2cor(self.WPLat.text())
                sheetWaypoints.cell(i, 5).number_format = '-#.######'
                sheetWaypoints.cell(i, 6).value = self.longt2cor(self.WPLongt.text())
                sheetWaypoints.cell(i, 6).number_format = '-#.######'
                sheetWaypoints.cell(i, 7).value = self.WPNav.text()

    def searchwp(self):
        self.cur.execute("SELECT * FROM Waypoints WHERE Name = \'"+self.WPSearchEdit.text()+"\'")
        data = self.cur.fetchall()
        # 实例化列表模型，添加数据
        slm = QStringListModel()
        self.WPList = []
        for per in data:
            self.WPList.append(per[0].__str__())

        # 设置模型列表视图，加载数据列表
        slm.setStringList(self.WPList)

        # 设置列表视图的模型
        self.WPSearchResult.setModel(slm)
        # 单击触发自定义的槽函数
        self.WPSearchResult.clicked.connect(self.selectwp)

    def selectwp(self):
        self.cur.execute("SELECT * FROM Waypoints WHERE ID = \'" + self.WPSearchResult.selectedIndexes()[0].data() + "\'")
        data = self.cur.fetchall()
        self.WPID.setText(data[0][0].__str__())
        self.WPIdent.setText(data[0][1].__str__())
        self.WPCollcated.setText(data[0][2].__str__())
        self.WPName.setText(data[0][3].__str__())
        self.WPLat.setText(self.cor2lat(data[0][4].__str__()))
        self.WPLongt.setText(self.cor2longt(data[0][5].__str__()))
        if data[0][6] != None:
            self.WPNav.setText(data[0][6].__str__())
        self.cur.execute("SELECT * FROM Waypointlookup WHERE ID = \'" + self.WPSearchResult.selectedIndexes()[0].data() + "\'")
        data = self.cur.fetchall()
        self.WPCountry.setText(data[0][1].__str__())

    def searchAirport(self):
        self.selectairport(self.AirportSearch.text())

    def selectairport(self,ICAO):
        # print()

        if ICAO.__str__().__len__()==4:
            ICAO = ICAO.upper()
            self.cur.execute("SELECT * FROM Airports WHERE ICAO=\'" + ICAO + "\' ")
        else:
            self.cur.execute("SELECT * FROM Airports WHERE ICAO=\'" + self.AirportSelect.selectedIndexes()[0].data() + "\' ")
        data = self.cur.fetchall()
        if data == []:
            return
        if data[0][0] !=None:
            self.AirportID.setText(data[0][0].__str__())
        if data[0][1] != None:
            self.AirportName.setText(data[0][1].__str__())
        if data[0][2] != None:
            self.AirportICAO.setText(data[0][2].__str__())
        if data[0][4] != None:
            self.AirportLat.setText(self.cor2lat(data[0][4]))
        if data[0][5] != None:
            self.AirportLongt.setText(self.cor2longt(data[0][5]))
        if data[0][6] != None:
            self.AirportElv.setText(data[0][6].__str__())
        if data[0][7] != None:
            self.AirportTA.setText(data[0][7].__str__())
        if data[0][8] != None:
            self.AirportTL.setText(data[0][8].__str__())
        if data[0][9] != None:
            self.AirportSpd.setText(data[0][9].__str__())
        if data[0][10] != None:
            self.AirportSpdH.setText(data[0][10].__str__())
        self.cur.execute("SELECT * FROM Runways WHERE AirportID=\'" + data[0][0].__str__() + "\'")

        data = self.cur.fetchall()
        self.RunwayList = []
        for per in data:
            self.RunwayList.append(per[2])
            # 实例化列表模型，添加数据
        run = QStringListModel()
        run.setStringList(self.RunwayList)
        # 设置列表视图的模型
        self.RunwaySelect.setModel(run)
        # 单击触发自定义的槽函数
        self.RunwaySelect.clicked.connect(self.selectrunways)
        if ICAO.__str__().__len__()==4:
            self.cur.execute("SELECT * FROM AirportCommunication WHERE airport_identifier=\'" + ICAO + "\' ")
        else:
            self.cur.execute("SELECT * FROM AirportCommunication WHERE airport_identifier=\'" + self.AirportSelect.selectedIndexes()[0].data() + "\' ")

        data = self.cur.fetchall()
        if data != []:
            self.AirportAICAO.setText(data[0][1].__str__())

        # print(self.AirportSelect.selectedIndexes()[0].data())

    def selectrunways(self):

        self.cur.execute("SELECT * FROM Runways WHERE AirportID =\'" + self.AirportID.text() + "\' AND Ident =\'"+self.RunwaySelect.selectedIndexes()[0].data()+"\'" )
        data = self.cur.fetchall()
        self.RunwayID.setText(data[0][0].__str__())
        self.RAirportID.setText(data[0][1].__str__())
        self.RIdent.setText(data[0][2].__str__())
        self.RTureHeading.setText(data[0][3].__str__())
        self.RLength.setText(data[0][4].__str__())
        self.RWidth.setText(data[0][5].__str__())
        self.RSurface.setCurrentText(data[0][6].__str__())
        self.RLat.setText(self.cor2lat(data[0][7].__str__()))
        self.RLongt.setText(self.cor2longt(data[0][8].__str__()))
        self.RHeight.setText(data[0][9].__str__())
        self.cur.execute("SELECT * FROM ILSes WHERE RunwayID =\'" + data[0][0].__str__() + "\' ")
        data = self.cur.fetchall()
        self.ILSCategory.clear()
        self.ILSCategory.addItem('I')
        self.ILSCategory.addItem('II')
        self.ILSCategory.addItem('III')
        self.ILSID.setText('')
        self.ILSFreq.setText('')
        self.ILSAngle.setText('')
        self.ILSLat.setText('')
        self.ILSLongt.setText('')
        self.ILSCategory.setCurrentText('')
        self.ILSIdent.setText('')
        self.LocHeading.setText('')
        self.ILSTCH.setText('')
        self.ILSDME.setChecked(0)
        self.ILSHeight.setText('')
        if data!=[]:
            self.ILSID.setText(data[0][0].__str__())
            IFreq = hex(int(data[0][2].__str__())).__str__()
            IFreq = IFreq[2:8]
            IFreq = (int(IFreq)*1.0/1000).__str__()
            self.ILSFreq.setText(IFreq)
            self.ILSAngle.setText(data[0][3].__str__())
            self.ILSLat.setText(self.cor2lat(data[0][4].__str__()))
            self.ILSLongt.setText(self.cor2longt(data[0][5].__str__()))
            if data[0][6].__str__() =='1':
                self.ILSCategory.setCurrentText('I')
            else:
                if data[0][6].__str__() =='2':
                    self.ILSCategory.setCurrentText('II')
                else:
                    if data[0][6].__str__() =='3':
                        self.ILSCategory.setCurrentText('III')
            self.ILSIdent.setText(data[0][7].__str__())
            self.LocHeading.setText(data[0][8].__str__())
            self.ILSTCH.setText(data[0][9].__str__())
            if data[0][10].__str__() =='1':
                self.ILSDME.setChecked(True)
            else:
                self.ILSDME.setChecked(False)
            self.ILSHeight.setText(data[0][11].__str__())

    def saveairport(self):
        global sheetAirports
        global sheetAirportLookup
        global sheetAirportCommunication
        maxr = sheetAirports.max_row + 1
        for i in range(2, maxr + 1):
            if sheetAirports.cell(i, 1).value.__str__() == self.AirportID.text().__str__():
                print('发现重合,在第' + i.__str__() + '行')
                sheetAirports.cell(i, 1).value = self.AirportID.text()
                sheetAirports.cell(i, 2).value = self.AirportName.text()
                sheetAirports.cell(i, 3).value = self.AirportICAO.text()
                sheetAirports.cell(i, 5).number_format = '-#.######'
                sheetAirports.cell(i, 5).value = self.lat2cor(self.AirportLat.text())
                sheetAirports.cell(i, 6).number_format = '-#.######'
                sheetAirports.cell(i, 6).value = self.longt2cor(self.AirportLongt.text())
                sheetAirports.cell(i, 7).value = self.AirportElv.text()
                sheetAirports.cell(i, 8).value = self.AirportTA.text()
                sheetAirports.cell(i, 9).value = self.AirportTL.text()
                sheetAirports.cell(i, 10).value = self.AirportSpd.text()
                sheetAirports.cell(i, 11).value = self.AirportSpdH.text()
                break
            if i == maxr:
                print('没有重合,在第' + i.__str__() + '行')
                sheetAirports.cell(i, 1).value = self.AirportID.text()
                sheetAirports.cell(i, 2).value = self.AirportName.text()
                sheetAirports.cell(i, 3).value = self.AirportICAO.text()
                sheetAirports.cell(i, 5).number_format = '-#.######'
                sheetAirports.cell(i, 5).value =self.lat2cor(self.AirportLat.text())
                sheetAirports.cell(i, 6).number_format = '-#.######'
                sheetAirports.cell(i, 6).value = self.longt2cor(self.AirportLongt.text())
                sheetAirports.cell(i, 7).value = self.AirportElv.text()
                sheetAirports.cell(i, 8).value = self.AirportTA.text()
                sheetAirports.cell(i, 9).value = self.AirportTL.text()
                sheetAirports.cell(i, 10).value = self.AirportSpd.text()
                sheetAirports.cell(i, 11).value = self.AirportSpdH.text()

        maxr = sheetAirportLookup.max_row + 1
        for i in range(2, maxr + 1):
            if sheetAirportLookup.cell(i, 2).value.__str__() == self.AirportID.text().__str__():
                print('发现重合,在第' + i.__str__() + '行')
                sheetAirportLookup.cell(i, 1).value = self.AirportAICAO.text()+self.AirportICAO.text()
                sheetAirportLookup.cell(i, 2).value = self.AirportID.text()
                break
            if i == maxr:
                print('没有重合,在第' + i.__str__() + '行')
                sheetAirportLookup.cell(i, 1).value = self.AirportAICAO.text() + self.AirportICAO.text()
                sheetAirportLookup.cell(i, 2).value = self.AirportID.text()

    def saveexcel(self):
        global workbook
        workbook.save("Result.xlsx")

    def saverunway(self):
        global sheetRunways
        maxr = sheetRunways.max_row + 1
        for i in range(2, maxr + 1):
            if sheetRunways.cell(i, 1).value.__str__() == self.RunwayID.text().__str__():
                print('发现重合,在第' + i.__str__() + '行')
                sheetRunways.cell(i, 1).value = self.RunwayID.text()
                sheetRunways.cell(i, 2).value = self.RAirportID.text()
                sheetRunways.cell(i, 3).value = self.RIdent.text()
                sheetRunways.cell(i, 4).value = self.RTureHeading.text()
                sheetRunways.cell(i, 5).value = self.RLength.text()
                sheetRunways.cell(i, 6).value = self.RWidth.text()
                sheetRunways.cell(i, 7).value = self.RSurface.currentText()
                sheetRunways.cell(i, 8).number_format = '-#.######'
                sheetRunways.cell(i, 8).value = self.lat2cor(self.RLat.text())
                sheetRunways.cell(i, 9).number_format = '-#.######'
                sheetRunways.cell(i, 9).value = self.longt2cor(self.RLongt.text())

                sheetRunways.cell(i, 10).value = self.RHeight.text()
                break
            if i == maxr:
                print('没有重合,在第' + i.__str__() + '行')
                sheetRunways.cell(i, 1).value = self.RunwayID.text()
                sheetRunways.cell(i, 2).value = self.RAirportID.text()
                sheetRunways.cell(i, 3).value = self.RIdent.text()
                sheetRunways.cell(i, 4).value = self.RTureHeading.text()
                sheetRunways.cell(i, 5).value = self.RLength.text()
                sheetRunways.cell(i, 6).value = self.RWidth.text()
                sheetRunways.cell(i, 7).value = self.RSurface.currentText()
                sheetRunways.cell(i, 8).number_format = '-#.######'
                sheetRunways.cell(i, 8).value = self.lat2cor(self.RLat.text())
                sheetRunways.cell(i, 9).number_format = '-#.######'
                sheetRunways.cell(i, 9).value = self.longt2cor(self.RLongt.text())
                sheetRunways.cell(i, 10).value = self.RHeight.text()

    def saveils(self):
        global sheetILSes
        maxr = sheetILSes.max_row + 1
        for i in range(2, maxr + 1):
            if sheetILSes.cell(i, 1).value.__str__() == self.ILSID.text().__str__():
                print('发现重合,在第' + i.__str__() + '行')
                sheetILSes.cell(i, 1).value = self.ILSID.text()
                sheetILSes.cell(i, 2).value = self.RunwayID.text()
                iff = self.ILSFreq.text()
                print(int(int(float(iff) * 10000).__str__(), 16).__str__())
                sheetILSes.cell(i, 3).value = int(int(float(iff) * 10000).__str__(), 16).__str__()
                sheetILSes.cell(i, 4).value = self.ILSAngle.text()
                sheetILSes.cell(i, 5).value = self.lat2cor(self.ILSLat.text())
                sheetILSes.cell(i, 5).number_format = '-#.######'
                sheetILSes.cell(i, 6).value = self.longt2cor(self.ILSLongt.text())
                sheetILSes.cell(i, 6).number_format = '-#.######'
                sheetILSes.cell(i, 7).value = (self.ILSCategory.currentIndex()+1).__str__()
                sheetILSes.cell(i, 8).value = self.ILSIdent.text()
                sheetILSes.cell(i, 9).value = self.LocHeading.text()
                sheetILSes.cell(i, 10).value = self.ILSTCH.text()
                if self.ILSDME.isChecked():
                    sheetILSes.cell(i, 11).value = 1
                else:
                    sheetILSes.cell(i, 11).value = 0

                sheetILSes.cell(i, 12).value = self.ILSHeight.text()
                break
            if i == maxr:
                print('没有重合,在第' + i.__str__() + '行')
                sheetILSes.cell(i, 1).value = self.ILSID.text()
                sheetILSes.cell(i, 2).value = self.RunwayID.text()
                iff = self.ILSFreq.text()
                print(int(int(float(iff) * 10000).__str__(), 16).__str__())
                sheetILSes.cell(i, 3).value = int(int(float(iff) * 10000).__str__(), 16).__str__()
                sheetILSes.cell(i, 4).value = self.ILSAngle.text()
                sheetILSes.cell(i, 5).value = self.lat2cor(self.ILSLat.text())
                sheetILSes.cell(i, 5).number_format = '-#.######'
                sheetILSes.cell(i, 6).value = self.longt2cor(self.ILSLongt.text())
                sheetILSes.cell(i, 6).number_format = '-#.######'
                sheetILSes.cell(i, 7).value = self.ILSCategory.currentIndex() + 1
                sheetILSes.cell(i, 8).value = self.ILSIdent.text()
                sheetILSes.cell(i, 9).value = self.LocHeading.text()
                sheetILSes.cell(i, 10).value = self.ILSTCH.text()
                if self.ILSDME.isChecked():
                    sheetILSes.cell(i, 11).value = 1
                else:
                    sheetILSes.cell(i, 11).value = 0

                sheetILSes.cell(i, 12).value = self.ILSHeight.text()



if __name__ == "__main__":
    #print((int(hex(17854464)[2:8])/1000.0).__str__())

    #print(int(int(110.7*10000).__str__(),16))
    # con = sqlite3.connect('C:\\ProgramData\\Fenix\\Navdata\\nd.db3')
    # cur = con.cursor()
    # cur.execute("SELECT ID,ICAO FROM Airports")
    # airport = cur.fetchall()
    # 固定的，PyQt5程序都需要QApplication对象。sys.argv是命令行参数列表，确保程序可以双击运行
    app = QApplication(sys.argv)
    # 初始化
    myWin = MyMainForm()
    # 将窗口控件显示在屏幕上
    myWin.show()
    # 程序运行，sys.exit方法确保程序完整退出。
    sys.exit(app.exec_())
